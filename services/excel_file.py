import os
from openpyxl.styles import Font
from openpyxl import Workbook, load_workbook

def save_to_excel(filename, data):

    wf = Font(color='FFFFFF')

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = (['City', 'Temp.', 'Temp fells', 'Humidity', 'Wind speed', 'Date'])
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = wf

    data = data.values()
    ws.append(list(data))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = wf

    wb.save(filename)



# def create_excel():
#     values = [22,23,24,25,26]
#     wb = Workbook()
#     ws = wb.active
#     ws.title = 'Temps'
#
#     wf = Font(color='FFFFFF')
#
#     headers = (['No.', 'Temp.', 'Test'])
#     ws.append(headers)
#
#     for col in range(1, len(headers) + 1):
#         ws.cell(row=1, column=col).font = wf
#
#     for index, value in enumerate(values, start=1):
#         row_data = [index, value, f'test {index}']
#         ws.append(row_data)
#         row_num = index + 1
#         for col in range(1, len(row_data) + 1):
#             ws.cell(row=row_num, column=col).font = wf
#
#     wb.save(r'C:\Users\Bartosz\PycharmProjects\WeatherAPI\Test.xlsx')
#
#
# create_excel()